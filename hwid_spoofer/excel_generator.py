"""Excel report generation for HWID changes."""

import logging
from datetime import datetime
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Generates Excel reports for HWID changes."""

    def __init__(self, filename: str = "HWID_Cambios.xlsx"):
        """Initialize the report generator.
        
        Args:
            filename: Output Excel filename
        """
        self.filename = filename
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "HWID Changes"

    def _setup_styles(self):
        """Setup Excel styles."""
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def generate_report(self, changes: List[Dict]) -> bool:
        """Generate Excel report with HWID changes.
        
        Args:
            changes: List of change records
            
        Returns:
            True if successful
        """
        try:
            self._setup_styles()
            
            # Add title
            self.worksheet.merge_cells('A1:G1')
            title_cell = self.worksheet['A1']
            title_cell.value = "REPORTE DE CAMBIOS HWID"
            title_cell.font = self.title_font
            title_cell.alignment = self.center_alignment
            
            # Add timestamp
            self.worksheet.merge_cells('A2:G2')
            timestamp_cell = self.worksheet['A2']
            timestamp_cell.value = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            timestamp_cell.alignment = self.center_alignment
            
            # Headers
            headers = [
                'Fecha/Hora',
                'Componente',
                'Descripción',
                'Serial Original',
                'Serial Nuevo',
                'Estado',
                'Notas'
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = self.worksheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.center_alignment
                cell.border = self.border
            
            # Add data
            for row_num, change in enumerate(changes, 5):
                self.worksheet.cell(row=row_num, column=1).value = change.get('timestamp', '')
                self.worksheet.cell(row=row_num, column=2).value = change.get('component', '')
                self.worksheet.cell(row=row_num, column=3).value = change.get('description', '')
                self.worksheet.cell(row=row_num, column=4).value = change.get('old_serial', '')
                self.worksheet.cell(row=row_num, column=5).value = change.get('new_serial', '')
                self.worksheet.cell(row=row_num, column=6).value = change.get('status', '')
                self.worksheet.cell(row=row_num, column=7).value = change.get('notes', '')
                
                # Apply styles
                for col_num in range(1, 8):
                    cell = self.worksheet.cell(row=row_num, column=col_num)
                    cell.border = self.border
                    cell.alignment = self.wrap_alignment if col_num == 7 else self.center_alignment
            
            # Adjust column widths
            column_widths = [20, 20, 25, 25, 25, 15, 30]
            for col_num, width in enumerate(column_widths, 1):
                self.worksheet.column_dimensions[get_column_letter(col_num)].width = width
            
            # Save
            self.workbook.save(self.filename)
            logger.info(f"Excel report saved: {self.filename}")
            return True
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            return False

    def add_summary_sheet(self, summary: Dict) -> bool:
        """Add summary sheet to the report.
        
        Args:
            summary: Summary data
            
        Returns:
            True if successful
        """
        try:
            summary_sheet = self.workbook.create_sheet("Resumen")
            
            # Add summary data
            row = 1
            for key, value in summary.items():
                summary_sheet.cell(row=row, column=1).value = key
                summary_sheet.cell(row=row, column=2).value = value
                row += 1
            
            return True
        except Exception as e:
            logger.error(f"Error adding summary sheet: {e}")
            return False
